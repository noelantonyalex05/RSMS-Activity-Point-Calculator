"""
Rajagiri Activity Points Tracker
=================================

Logs into the RSMS student portal, sweeps every Class Code x Category
combination on the Activity Point Form, scrapes whatever submissions
exist for each, and writes everything to a nicely formatted Excel file -
including a Summary sheet with total points, a per-category breakdown,
and a pending-submissions list.

Confirmed site behavior:
  - Login page: studentlogin/login.php. Google sign-in requires a human,
    so you log in manually in the browser window this script opens.
  - Activity.asp has exactly two <select> elements before you interact
    with anything: [0] = Class Code, [1] = Category.
  - Clicking "Add Activity" is READ-ONLY - it reveals an entry form AND,
    below it, a results table (if any submissions exist for that
    class+category). It does NOT create a new record. The real
    record-creating action is a separate "SUBMIT" button inside the
    revealed form, which this script never touches.
  - Different categories have genuinely different table columns (e.g.
    Sports/Games has "Level"/"Points" where Leadership has "Documentary
    evidence"/"Rating By Faculty"). Columns are matched BY NAME so every
    value lands in the right place regardless of which category it came
    from.

Output: activity_points.xlsx, saved after every single combination so
progress is never lost if the script errors out or is interrupted.
  - Sheet "Summary"       - total points, category breakdown, pending list
  - Sheet "ActivityPoints" - every scraped row, one column per unique
                             field name encountered across all categories
  - Sheet "Skipped"        - only created if a combo failed after retries

Install once:
    pip install playwright openpyxl
    playwright install chromium
    (if 'playwright' isn't recognized as a command on Windows:
     python -m playwright install chromium)

Run:
    python activity_points_tracker.py
"""

import re
import time

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PWTimeoutError
from playwright.sync_api import sync_playwright

# ----------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------

LOGIN_URL = "https://rajagiritech.ac.in/stud/KTU/Student/studentlogin/login.php"
OUTPUT_FILE = "activity_points.xlsx"

DELAY_SECONDS = 1.0            # pause between combinations, be polite to the server
OPTIONS_WAIT_TIMEOUT = 15      # seconds to wait for a <select>'s options to populate
MAX_RETRIES_PER_COMBO = 2      # attempts before a combo is logged to Skipped

# Keyword -> bucket name, matched case-insensitively against the table's
# own "Category" column value (e.g. "Professional and Co-curricular").
BUCKET_KEYWORDS = [
    ("Professional", "professional"),
    ("Extracurricular", "extracurricular"),
    ("Leadership", "leadership"),
]

# ----------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBHEAD_FONT = Font(bold=True, size=12, color="1F4E78")
KPI_LABEL_FONT = Font(bold=True, size=11)
KPI_VALUE_FONT = Font(bold=True, size=14, color="2E7D32")

APPROVED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

_thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# Different categories label the same underlying field differently.
# Any header on the right maps to the canonical name on the left, so they
# share one Excel column instead of each spawning its own.
COLUMN_ALIASES = {
    "Organizing Institution / Society / Company": [
        "Name of professional society/ Association name/Name of Club etc.",
        "Name of the organizing instituition and Place",
        "Name of the offering agency",
        "Name of the Company and Address",
        "Organized By (name of institution with place)",
    ],
    "Evidence / Certificate": [
        "Documentary evidence (File Size<500kb)",
        "Certificate (File Size<500kb)",
        "Certificate/ Documentary evidence (File Size<500kb)",
        "Certificate/Letter from Authorities/Documentary evidence (File Size<500kb)",
    ],
}

_VARIANT_TO_CANONICAL = {
    variant: canonical
    for canonical, variants in COLUMN_ALIASES.items()
    for variant in variants
}


def canonical_field_name(name):
    """Map a scraped column header to its canonical name, if it's a known
    alias. Unknown headers pass through unchanged."""
    return _VARIANT_TO_CANONICAL.get(name, name)


def merge_field(row_dict, name, value):
    """Add (name, value) to row_dict under its canonical name. If that
    canonical field is already set in this row from a different alias
    (shouldn't normally happen - each category table only uses one
    variant - but handled just in case), the two values are concatenated
    instead of one silently overwriting the other."""
    canonical = canonical_field_name(name)
    existing = row_dict.get(canonical)
    if not existing:
        row_dict[canonical] = value
    elif value and value != existing:
        row_dict[canonical] = f"{existing}; {value}"




class ExcelBuilder:
    """Writes rows to a sheet, creating a new column the first time a
    field name is seen and reusing it after that - so different
    categories' differently-shaped tables all line up correctly."""

    def __init__(self, wb, sheet_name):
        self.wb = wb
        self.ws = wb.active
        self.ws.title = sheet_name
        self.header_to_col = {}
        self.next_col = 1
        self.widths = {}
        self.data_row = 1
        self.ws.freeze_panes = "A2"

    def get_col(self, name):
        if name not in self.header_to_col:
            col = self.next_col
            self.header_to_col[name] = col
            self._set_cell(1, col, name, header=True)
            self.next_col += 1
        return self.header_to_col[name]

    def _set_cell(self, row, col, value, header=False, fill=None):
        cell = self.ws.cell(row=row, column=col, value=value)
        if header:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        else:
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        length = len(str(value)) if value is not None else 0
        self.widths[col] = max(self.widths.get(col, 0), length)
        return cell

    def write_row(self, row_dict, fill=None):
        self.data_row += 1
        row = self.data_row
        for name, value in row_dict.items():
            col = self.get_col(name)
            self._set_cell(row, col, value, fill=fill)
        return row

    def autofit(self):
        for col, w in self.widths.items():
            letter = get_column_letter(col)
            self.ws.column_dimensions[letter].width = min(max(w + 2, 10), 60)


def write_skipped_sheet(wb, skipped_rows):
    """(Re)writes the Skipped sheet from scratch. No-op if nothing failed."""
    if "Skipped" in wb.sheetnames:
        del wb["Skipped"]
    if not skipped_rows:
        return
    ws = wb.create_sheet("Skipped")
    headers = ["Class Code", "Category", "Error"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    for ri, row in enumerate(skipped_rows, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v).border = THIN_BORDER
    for ci in range(1, 4):
        ws.column_dimensions[get_column_letter(ci)].width = 34


# ----------------------------------------------------------------------
# Summary calculation
# ----------------------------------------------------------------------

def bucket_for(category_value):
    if not category_value:
        return "Other / Unclassified"
    lower = str(category_value).lower()
    for bucket_name, keyword in BUCKET_KEYWORDS:
        if keyword in lower:
            return bucket_name
    return "Other / Unclassified"


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s == "-":
        return None
    m = re.search(r"-?\d+(\.\d+)?", s)
    if not m:
        return None
    return float(m.group())


def find_points_value(row_dict):
    """Sum any field whose name contains 'point' or 'rating' (excluding
    the 'Point Status' text field) - this is where the numeric point
    value lives, under whatever name that category's table uses for it."""
    total = 0.0
    found_any = False
    for name, value in row_dict.items():
        name_l = name.lower()
        if name_l == "point status":
            continue
        if "point" in name_l or "rating" in name_l:
            n = to_number(value)
            if n is not None:
                total += n
                found_any = True
    return total if found_any else 0.0


class SummaryStats:
    def __init__(self):
        self.bucket_totals = {}
        self.bucket_counts = {}
        self.bucket_pending = {}
        self.total_approved_points = 0.0
        self.total_approved_count = 0
        self.pending_rows = []
        self.approved_rows = []

    def accumulate(self, row_dict):
        """Update running totals from one scraped row. Returns the row's
        status ('approved' / 'pending' / other, lowercased) so the caller
        can color the row accordingly."""
        status = str(row_dict.get("Point Status", "") or "").strip().lower()
        category_value = row_dict.get("Category")
        bucket = bucket_for(category_value)
        points = find_points_value(row_dict)

        if status == "approved":
            self.bucket_totals[bucket] = self.bucket_totals.get(bucket, 0.0) + points
            self.bucket_counts[bucket] = self.bucket_counts.get(bucket, 0) + 1
            self.total_approved_points += points
            self.total_approved_count += 1
            self.approved_rows.append({
                "Class Code": row_dict.get("Class Code"),
                "Selected Category": row_dict.get("Selected Category"),
                "Category": category_value,
                "Activity": row_dict.get("Activity"),
                "Name of event": row_dict.get("Name of event"),
                "Points": points,
            })
        elif status == "pending":
            self.bucket_pending[bucket] = self.bucket_pending.get(bucket, 0) + 1
            self.pending_rows.append({
                "Class Code": row_dict.get("Class Code"),
                "Selected Category": row_dict.get("Selected Category"),
                "Category": category_value,
                "Activity": row_dict.get("Activity"),
                "Name of event": row_dict.get("Name of event"),
            })
        return status

    def ordered_buckets(self):
        all_buckets = set(self.bucket_totals) | set(self.bucket_counts) | set(self.bucket_pending)
        ordered = [b for b, _ in BUCKET_KEYWORDS if b in all_buckets]
        ordered += sorted(b for b in all_buckets if b not in ordered)
        return ordered

    def write_sheet(self, wb):
        """(Re)builds the Summary sheet from current totals, as the first tab."""
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws = wb.create_sheet("Summary", 0)

        ws.merge_cells("A1:D1")
        title = ws.cell(row=1, column=1, value="Activity Points Summary")
        title.font = TITLE_FONT
        ws.row_dimensions[1].height = 26

        r = 3
        ws.cell(row=r, column=1, value="Total Approved Points").font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=self.total_approved_points).font = KPI_VALUE_FONT
        r += 1
        ws.cell(row=r, column=1, value="Total Approved Submissions").font = KPI_LABEL_FONT
        ws.cell(row=r, column=2, value=self.total_approved_count).font = Font(bold=True, size=12)
        r += 1
        ws.cell(row=r, column=1, value="Pending Submissions").font = KPI_LABEL_FONT
        pending_cell = ws.cell(row=r, column=2, value=len(self.pending_rows))
        pending_cell.font = Font(bold=True, size=12, color="BF8F00")
        r += 2

        ws.cell(row=r, column=1, value="Points by Category").font = SUBHEAD_FONT
        r += 1
        headers = ["Category", "Approved Points", "Approved Count", "Pending Count"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        r += 1
        for b in self.ordered_buckets():
            ws.cell(row=r, column=1, value=b).border = THIN_BORDER
            ws.cell(row=r, column=2, value=self.bucket_totals.get(b, 0.0)).border = THIN_BORDER
            ws.cell(row=r, column=3, value=self.bucket_counts.get(b, 0)).border = THIN_BORDER
            ws.cell(row=r, column=4, value=self.bucket_pending.get(b, 0)).border = THIN_BORDER
            r += 1
        r += 1

        ws.cell(row=r, column=1,
                value=f"Approved Submissions ({len(self.approved_rows)})").font = SUBHEAD_FONT
        r += 1
        aheaders = ["Class Code", "Selected Category", "Category", "Activity", "Name of event", "Points"]
        for ci, h in enumerate(aheaders, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        r += 1
        for ar in self.approved_rows:
            for ci, key in enumerate(aheaders, 1):
                cell = ws.cell(row=r, column=ci, value=ar.get(key))
                cell.fill = APPROVED_FILL
                cell.border = THIN_BORDER
            r += 1
        r += 1

        ws.cell(row=r, column=1,
                value=f"Pending Submissions ({len(self.pending_rows)})").font = SUBHEAD_FONT
        r += 1
        pheaders = ["Class Code", "Selected Category", "Category", "Activity", "Name of event"]
        for ci, h in enumerate(pheaders, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        r += 1
        for pr in self.pending_rows:
            for ci, key in enumerate(pheaders, 1):
                cell = ws.cell(row=r, column=ci, value=pr.get(key))
                cell.fill = PENDING_FILL
                cell.border = THIN_BORDER
            r += 1

        widths = [26, 34, 34, 26, 30, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------------------------
# Playwright automation
# ----------------------------------------------------------------------

def get_select_options(select_locator):
    """Return list of (value, label) for every real option in a <select>."""
    options = select_locator.locator("option").all()
    result = []
    for opt in options:
        value = opt.get_attribute("value")
        label = opt.inner_text().strip()
        if value is None or value == "":
            continue
        result.append((value, label))
    return result


def wait_for_options_populated(select_locator, min_options=1, timeout_s=OPTIONS_WAIT_TIMEOUT):
    """Poll until the <select> has more than min_options options, or raise."""
    deadline = time.time() + timeout_s
    count = 0
    while time.time() < deadline:
        count = select_locator.locator("option").count()
        if count > min_options:
            return
        time.sleep(0.2)
    raise TimeoutError(
        f"Select did not populate options within {timeout_s}s (had {count})"
    )


def scrape_one_combo(page, form_url, class_val, cat_val):
    """Navigate to the form, select both dropdowns, click Add Activity
    (read-only - never clicks SUBMIT), and return (header_cells, rows)."""
    page.goto(form_url)
    page.wait_for_load_state("load")

    class_select = page.locator("select").nth(0)
    category_select = page.locator("select").nth(1)

    wait_for_options_populated(class_select)
    wait_for_options_populated(category_select)

    class_select.select_option(class_val, timeout=10000)
    category_select.select_option(cat_val, timeout=10000)

    page.click("text=Add Activity")
    page.wait_for_load_state("load")
    time.sleep(0.5)  # small buffer for classic ASP postback rendering

    results_table = page.locator("table", has_text="Sl.No")
    if results_table.count() == 0:
        return None, []

    rows = results_table.first.locator("tr").all()
    header_cells = None
    scraped = []
    for row in rows:
        cells = [c.inner_text().strip() for c in row.locator("td, th").all()]
        if not cells:
            continue
        if cells[0] == "Sl.No":
            if header_cells is None:
                header_cells = cells
            continue  # header row, not data
        scraped.append(cells)
    return header_cells, scraped


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()
    builder = ExcelBuilder(wb, "ActivityPoints")
    builder.get_col("Class Code")
    builder.get_col("Selected Category")

    stats = SummaryStats()
    skipped_rows = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto(LOGIN_URL)

        print(">>> Please log in with Google manually.")
        page.wait_for_url("**/Home.asp", timeout=0)
        print(">>> Login detected. Navigating to Activity Point Form...")

        page.click("text=Activity Point Form")
        page.wait_for_load_state("load")
        form_url = page.url

        class_select = page.locator("select").nth(0)
        category_select = page.locator("select").nth(1)
        wait_for_options_populated(class_select)
        wait_for_options_populated(category_select)

        class_options = get_select_options(class_select)
        category_options = get_select_options(category_select)

        print(f">>> Found {len(class_options)} class codes, "
              f"{len(category_options)} categories "
              f"({len(class_options) * len(category_options)} combinations total)")

        combos = [(c, a) for c in class_options for a in category_options]

        for i, ((class_val, class_label), (cat_val, cat_label)) in enumerate(combos, 1):
            last_error = None
            header_cells = None
            scraped_rows = None

            for attempt in range(1, MAX_RETRIES_PER_COMBO + 1):
                try:
                    header_cells, scraped_rows = scrape_one_combo(
                        page, form_url, class_val, cat_val
                    )
                    last_error = None
                    break
                except (PWTimeoutError, TimeoutError, Exception) as e:
                    last_error = e
                    print(f"    [{i}/{len(combos)}] class={class_label} "
                          f"category={cat_label} -> attempt {attempt} failed: {e}")
                    time.sleep(1.5)

            if last_error is not None:
                print(f"    [{i}/{len(combos)}] SKIPPING class={class_label} "
                      f"category={cat_label} after {MAX_RETRIES_PER_COMBO} attempts")
                skipped_rows.append([class_label, cat_label, str(last_error)])
                write_skipped_sheet(wb, skipped_rows)
                wb.save(OUTPUT_FILE)
                time.sleep(DELAY_SECONDS)
                continue

            if not scraped_rows:
                print(f"    [{i}/{len(combos)}] class={class_label} "
                      f"category={cat_label} -> no submissions")
            else:
                for cells in scraped_rows:
                    row_dict = {"Class Code": class_label, "Selected Category": cat_label}
                    if header_cells:
                        for name, value in zip(header_cells, cells):
                            merge_field(row_dict, name, value)
                    else:
                        for idx, value in enumerate(cells, 1):
                            merge_field(row_dict, f"Col{idx}", value)

                    status = stats.accumulate(row_dict)
                    fill = APPROVED_FILL if status == "approved" else (
                        PENDING_FILL if status == "pending" else None
                    )
                    builder.write_row(row_dict, fill=fill)

                print(f"    [{i}/{len(combos)}] class={class_label} "
                      f"category={cat_label} -> {len(scraped_rows)} row(s)")

            builder.autofit()
            stats.write_sheet(wb)
            wb.save(OUTPUT_FILE)
            time.sleep(DELAY_SECONDS)

        builder.autofit()
        stats.write_sheet(wb)
        wb.save(OUTPUT_FILE)
        print(f"\n>>> Done. Total approved points: {stats.total_approved_points}")
        print(f">>> Pending submissions: {len(stats.pending_rows)}")
        print(f">>> Saved to {OUTPUT_FILE}")
        browser.close()


if __name__ == "__main__":
    main()
